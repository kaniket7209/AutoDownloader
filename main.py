import pandas as pd
import requests
import tabula


# p.save_book_as(file_name='gst_V0FnyZ_2022-05-15.xls',
#                dest_file_name='mainFile.xlsx')

def to_excel(file):
    df = pd.read_excel(file, header=None)
    print(df)
    df.to_csv('data.csv', index=False, header=False)
    df.to_excel('output.xlsx', index=False, header=False)

import openpyxl
# df = pd.read_excel('gst_V0FnyZ_2022-05-15.xls', engine='openpyxl')
def meeshoPdf(file,loc):
# book = openpyxl.load_workbook('test1.xlsx',data_only=True)
    book = openpyxl.load_workbook(file,data_only=True)
    sheet = book.active

    len = sheet.max_row
    i = 4
    while(i < len):
        a = sheet.cell(row=i, column=31)
        i= i +1
        print(a.value,i)
        s = requests.get("{}".format(a.value))
        # pdf = open("Files/pdf"+str(i)+".pdf", 'wb')
        pdf = open(loc+"/pdf"+str(i)+".pdf", 'wb')
        
        pdf.write(s.content)
        pdf.close()
        print("File ", i, " downloaded")

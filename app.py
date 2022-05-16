import pandas as pd
import requests
import openpyxl
import tabula
from tabula import read_pdf
def meeshoPdf(file,loc):

    book = openpyxl.load_workbook(file,data_only=True)
    # book = openpyxl.load_workbook('test1.xlsx',data_only=True)
    sheet = book.active

    len = sheet.max_row
    i = 4
    while(i < len):
        a = sheet.cell(row=i, column=31)
        i= i +1
        print(a.value,i)
        s = requests.get("{}".format(a.value))
        pdf = open(loc+str(i)+".pdf", 'wb')
        # pdf = open("Files/pdf"+str(i)+".pdf", 'wb')
        
        pdf.write(s.content)
        pdf.close()
        print("File ", i, " downloaded")
        # df = read_pdf("Files/pdf{}.pdf".format(i), pages='all')[0]
        df = tabula.read_pdf("Files/pdf{}.pdf".format(i), output_format='json', pages='all', lattice=False)[0]
        # convert PDF into CSV
        # tabula.convert_into("Files/pdf{}.pdf".format(i), "data.csv", output_format="csv", pages='all')
        print(df)
